"""
Sistema de Exportação de Relatórios
Suporta CSV, Excel e PDF com formatação avançada
"""

import io
import csv
import json
import pandas as pd
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Union
from pathlib import Path
import asyncio
from sqlalchemy import select, and_, or_, func, text
from sqlalchemy.ext.asyncio import AsyncSession

# ReportLab para PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel/CSV
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows

from app.database import AsyncSessionLocal
from app.models.database import (
    Appointment, Conversation, AdminUser, 
    PushSubscription, PushNotification
)

class ReportExportService:
    """Serviço principal para exportação de relatórios"""
    
    def __init__(self):
        self.temp_dir = Path("temp_reports")
        self.temp_dir.mkdir(exist_ok=True)
        
        # Configurações de estilo
        self.colors = {
            'primary': '#366092',
            'secondary': '#4a7cb8',
            'success': '#10b981',
            'warning': '#f59e0b',
            'error': '#ef4444',
            'gray': '#6b7280'
        }
        
        self.fonts = {
            'title': ('Helvetica-Bold', 16),
            'subtitle': ('Helvetica-Bold', 12),
            'body': ('Helvetica', 10),
            'small': ('Helvetica', 8)
        }

    async def export_appointments_report(
        self, 
        format_type: str = 'excel',
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        status: Optional[str] = None,
        user_id: Optional[int] = None
    ) -> tuple[bytes, str]:
        """Exportar relatório de agendamentos"""
        
        # Buscar dados
        async with AsyncSessionLocal() as session:
            query = select(Appointment)
            
            # Filtros
            conditions = []
            if date_from:
                conditions.append(func.date(Appointment.date_time) >= date_from)
            if date_to:
                conditions.append(func.date(Appointment.date_time) <= date_to)
            if status:
                conditions.append(Appointment.status == status)
            if user_id:
                conditions.append(Appointment.user_id == user_id)
            
            if conditions:
                query = query.where(and_(*conditions))
            
            query = query.order_by(Appointment.date_time.desc())
            result = await session.execute(query)
            appointments = result.scalars().all()
        
        # Converter para dados estruturados
        data = []
        for apt in appointments:
            data.append({
                'ID': apt.id,
                'Data/Hora': apt.date_time.strftime('%d/%m/%Y %H:%M'),
                'Cliente': apt.user_name or 'N/A',
                'Telefone': apt.user_phone or 'N/A',
                'Status': self._translate_status(apt.status),
                'Serviço': apt.service_type or 'Geral',
                'Observações': apt.notes or '',
                'Criado em': apt.created_at.strftime('%d/%m/%Y %H:%M') if apt.created_at else '',
                'Atualizado em': apt.updated_at.strftime('%d/%m/%Y %H:%M') if apt.updated_at else ''
            })
        
        # Dados para resumo
        summary_data = {
            'total_appointments': len(data),
            'period': f"{date_from or 'Início'} até {date_to or 'Hoje'}",
            'status_breakdown': self._calculate_status_breakdown([apt.status for apt in appointments]),
            'generated_at': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        }
        
        # Exportar no formato solicitado
        if format_type.lower() == 'csv':
            return await self._export_csv(data, 'agendamentos'), 'text/csv'
        elif format_type.lower() == 'excel':
            return await self._export_excel_appointments(data, summary_data), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif format_type.lower() == 'pdf':
            return await self._export_pdf_appointments(data, summary_data), 'application/pdf'
        else:
            raise ValueError(f"Formato não suportado: {format_type}")

    async def export_conversations_report(
        self,
        format_type: str = 'excel',
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        user_id: Optional[int] = None
    ) -> tuple[bytes, str]:
        """Exportar relatório de conversas"""
        
        async with AsyncSessionLocal() as session:
            query = select(Conversation)
            
            conditions = []
            if date_from:
                conditions.append(func.date(Conversation.created_at) >= date_from)
            if date_to:
                conditions.append(func.date(Conversation.created_at) <= date_to)
            if user_id:
                conditions.append(Conversation.user_id == user_id)
            
            if conditions:
                query = query.where(and_(*conditions))
            
            query = query.order_by(Conversation.created_at.desc())
            result = await session.execute(query)
            conversations = result.scalars().all()
        
        data = []
        for conv in conversations:
            data.append({
                'ID': conv.id,
                'Usuário': conv.user_name or f'User {conv.user_id}',
                'Telefone': conv.user_phone or 'N/A',
                'Última Mensagem': conv.last_message or 'N/A',
                'Status': self._translate_conversation_status(conv.status),
                'Criado em': conv.created_at.strftime('%d/%m/%Y %H:%M') if conv.created_at else '',
                'Atualizado em': conv.updated_at.strftime('%d/%m/%Y %H:%M') if conv.updated_at else '',
                'Contexto': conv.context or ''
            })
        
        summary_data = {
            'total_conversations': len(data),
            'period': f"{date_from or 'Início'} até {date_to or 'Hoje'}",
            'status_breakdown': self._calculate_conversation_status_breakdown([conv.status for conv in conversations]),
            'generated_at': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        }
        
        if format_type.lower() == 'csv':
            return await self._export_csv(data, 'conversas'), 'text/csv'
        elif format_type.lower() == 'excel':
            return await self._export_excel_conversations(data, summary_data), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif format_type.lower() == 'pdf':
            return await self._export_pdf_conversations(data, summary_data), 'application/pdf'
        else:
            raise ValueError(f"Formato não suportado: {format_type}")

    async def export_dashboard_report(
        self,
        format_type: str = 'excel',
        date_from: Optional[date] = None,
        date_to: Optional[date] = None
    ) -> tuple[bytes, str]:
        """Exportar relatório de dashboard com métricas"""
        
        async with AsyncSessionLocal() as session:
            # Métricas de agendamentos
            apt_query = select(
                func.count(Appointment.id).label('total_appointments'),
                func.count().filter(Appointment.status == 'pending').label('pending_appointments'),
                func.count().filter(Appointment.status == 'confirmed').label('confirmed_appointments'),
                func.count().filter(Appointment.status == 'completed').label('completed_appointments'),
                func.count().filter(Appointment.status == 'cancelled').label('cancelled_appointments')
            )
            
            if date_from:
                apt_query = apt_query.where(func.date(Appointment.date_time) >= date_from)
            if date_to:
                apt_query = apt_query.where(func.date(Appointment.date_time) <= date_to)
            
            apt_result = await session.execute(apt_query)
            apt_metrics = apt_result.first()
            
            # Métricas de conversas
            conv_query = select(
                func.count(Conversation.id).label('total_conversations'),
                func.count().filter(Conversation.status == 'active').label('active_conversations'),
                func.count().filter(Conversation.status == 'closed').label('closed_conversations')
            )
            
            if date_from:
                conv_query = conv_query.where(func.date(Conversation.created_at) >= date_from)
            if date_to:
                conv_query = conv_query.where(func.date(Conversation.created_at) <= date_to)
            
            conv_result = await session.execute(conv_query)
            conv_metrics = conv_result.first()
            
            # Métricas de push notifications
            push_query = select(
                func.count(PushNotification.id).label('total_notifications')
            )
            
            if date_from:
                push_query = push_query.where(func.date(PushNotification.sent_at) >= date_from)
            if date_to:
                push_query = push_query.where(func.date(PushNotification.sent_at) <= date_to)
            
            push_result = await session.execute(push_query)
            push_metrics = push_result.first()
        
        # Consolidar métricas
        metrics = {
            'período': f"{date_from or 'Início'} até {date_to or 'Hoje'}",
            'total_agendamentos': apt_metrics.total_appointments,
            'agendamentos_pendentes': apt_metrics.pending_appointments,
            'agendamentos_confirmados': apt_metrics.confirmed_appointments,
            'agendamentos_concluídos': apt_metrics.completed_appointments,
            'agendamentos_cancelados': apt_metrics.cancelled_appointments,
            'total_conversas': conv_metrics.total_conversations,
            'conversas_ativas': conv_metrics.active_conversations,
            'conversas_encerradas': conv_metrics.closed_conversations,
            'notificações_enviadas': push_metrics.total_notifications,
            'gerado_em': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        }
        
        # Dados por dia (últimos 30 dias)
        daily_data = await self._get_daily_metrics(date_from, date_to)
        
        if format_type.lower() == 'excel':
            return await self._export_excel_dashboard(metrics, daily_data), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif format_type.lower() == 'pdf':
            return await self._export_pdf_dashboard(metrics, daily_data), 'application/pdf'
        else:
            # CSV simples das métricas
            csv_data = [{'Métrica': k, 'Valor': v} for k, v in metrics.items()]
            return await self._export_csv(csv_data, 'dashboard'), 'text/csv'

    # Métodos de exportação específicos
    async def _export_csv(self, data: List[Dict], filename_prefix: str) -> bytes:
        """Exportar dados para CSV"""
        if not data:
            return b"Nenhum dado encontrado"
        
        output = io.StringIO()
        fieldnames = data[0].keys()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue().encode('utf-8-sig')  # BOM para Excel brasileiro

    async def _export_excel_appointments(self, data: List[Dict], summary: Dict) -> bytes:
        """Exportar agendamentos para Excel com formatação"""
        wb = openpyxl.Workbook()
        
        # Aba Resumo
        summary_ws = wb.active
        summary_ws.title = "Resumo"
        
        # Header do resumo
        summary_ws['A1'] = "RELATÓRIO DE AGENDAMENTOS"
        summary_ws['A1'].font = Font(size=16, bold=True)
        summary_ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        summary_ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        summary_ws.merge_cells('A1:D1')
        
        # Informações do resumo
        row = 3
        for key, value in summary.items():
            summary_ws[f'A{row}'] = key.replace('_', ' ').title()
            summary_ws[f'B{row}'] = str(value)
            summary_ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Aba de Dados
        data_ws = wb.create_sheet("Dados")
        
        if data:
            # Converter para DataFrame para facilitar
            df = pd.DataFrame(data)
            
            # Escrever headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = data_ws.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4a7cb8", end_color="4a7cb8", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Escrever dados
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    data_ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Ajustar largura das colunas
            for column in data_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = min(max_length + 2, 50)
                data_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def _export_excel_conversations(self, data: List[Dict], summary: Dict) -> bytes:
        """Exportar conversas para Excel"""
        wb = openpyxl.Workbook()
        
        # Similar ao appointments mas com dados de conversas
        summary_ws = wb.active
        summary_ws.title = "Resumo"
        
        summary_ws['A1'] = "RELATÓRIO DE CONVERSAS"
        summary_ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        summary_ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        summary_ws.merge_cells('A1:D1')
        
        row = 3
        for key, value in summary.items():
            summary_ws[f'A{row}'] = key.replace('_', ' ').title()
            summary_ws[f'B{row}'] = str(value)
            summary_ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Dados
        data_ws = wb.create_sheet("Conversas")
        if data:
            df = pd.DataFrame(data)
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = data_ws.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
            
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    data_ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Ajustar colunas
            for column in data_ws.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                data_ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def _export_excel_dashboard(self, metrics: Dict, daily_data: List[Dict]) -> bytes:
        """Exportar dashboard para Excel com gráficos"""
        wb = openpyxl.Workbook()
        
        # Aba Métricas
        metrics_ws = wb.active
        metrics_ws.title = "Métricas"
        
        metrics_ws['A1'] = "DASHBOARD - MÉTRICAS GERAIS"
        metrics_ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        metrics_ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        metrics_ws.merge_cells('A1:C1')
        
        row = 3
        for key, value in metrics.items():
            metrics_ws[f'A{row}'] = key.replace('_', ' ').title()
            metrics_ws[f'B{row}'] = value
            metrics_ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Aba de Dados Diários (se disponível)
        if daily_data:
            daily_ws = wb.create_sheet("Dados Diários")
            df_daily = pd.DataFrame(daily_data)
            
            # Headers
            for col_num, column_title in enumerate(df_daily.columns, 1):
                cell = daily_ws.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
            
            # Dados
            for row_num, row_data in enumerate(df_daily.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    daily_ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Criar gráfico se há dados suficientes
            if len(daily_data) > 1:
                chart = LineChart()
                chart.title = "Tendência Diária - Agendamentos"
                chart.y_axis.title = 'Quantidade'
                chart.x_axis.title = 'Data'
                
                # Dados para o gráfico (assumindo colunas data e agendamentos)
                data_ref = Reference(daily_ws, min_col=2, min_row=1, max_row=len(daily_data)+1, max_col=2)
                cats_ref = Reference(daily_ws, min_col=1, min_row=2, max_row=len(daily_data)+1)
                
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                
                daily_ws.add_chart(chart, "E2")
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def _export_pdf_appointments(self, data: List[Dict], summary: Dict) -> bytes:
        """Exportar agendamentos para PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("RELATÓRIO DE AGENDAMENTOS", title_style))
        
        # Resumo
        elements.append(Paragraph("RESUMO EXECUTIVO", styles['Heading2']))
        
        for key, value in summary.items():
            if key != 'status_breakdown':
                text = f"<b>{key.replace('_', ' ').title()}:</b> {value}"
                elements.append(Paragraph(text, styles['Normal']))
        
        elements.append(Spacer(1, 20))
        
        # Tabela de dados (primeiras 50 linhas para evitar PDF muito grande)
        if data:
            elements.append(Paragraph("DETALHAMENTO DOS AGENDAMENTOS", styles['Heading2']))
            
            # Preparar dados para tabela
            table_data = []
            headers = ['ID', 'Data/Hora', 'Cliente', 'Status', 'Serviço']
            table_data.append(headers)
            
            for item in data[:50]:  # Limitar para performance
                row = [
                    str(item.get('ID', '')),
                    str(item.get('Data/Hora', '')),
                    str(item.get('Cliente', ''))[:20],  # Truncar nomes longos
                    str(item.get('Status', '')),
                    str(item.get('Serviço', ''))[:15]
                ]
                table_data.append(row)
            
            # Criar tabela
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            
            if len(data) > 50:
                elements.append(Spacer(1, 20))
                elements.append(Paragraph(f"<i>Mostrando primeiros 50 de {len(data)} registros</i>", styles['Normal']))
        
        # Rodapé
        elements.append(Spacer(1, 30))
        footer_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}"
        elements.append(Paragraph(footer_text, styles['Normal']))
        
        doc.build(elements)
        return buffer.getvalue()

    async def _export_pdf_conversations(self, data: List[Dict], summary: Dict) -> bytes:
        """Exportar conversas para PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#10b981'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("RELATÓRIO DE CONVERSAS", title_style))
        
        # Resumo
        elements.append(Paragraph("RESUMO EXECUTIVO", styles['Heading2']))
        for key, value in summary.items():
            if key != 'status_breakdown':
                text = f"<b>{key.replace('_', ' ').title()}:</b> {value}"
                elements.append(Paragraph(text, styles['Normal']))
        
        elements.append(Spacer(1, 20))
        
        # Dados
        if data:
            elements.append(Paragraph("DETALHAMENTO DAS CONVERSAS", styles['Heading2']))
            
            table_data = [['ID', 'Usuário', 'Telefone', 'Status', 'Última Msg']]
            
            for item in data[:40]:
                row = [
                    str(item.get('ID', '')),
                    str(item.get('Usuário', ''))[:15],
                    str(item.get('Telefone', '')),
                    str(item.get('Status', '')),
                    str(item.get('Última Mensagem', ''))[:25]
                ]
                table_data.append(row)
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        
        footer_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}"
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(footer_text, styles['Normal']))
        
        doc.build(elements)
        return buffer.getvalue()

    async def _export_pdf_dashboard(self, metrics: Dict, daily_data: List[Dict]) -> bytes:
        """Exportar dashboard para PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("DASHBOARD - RELATÓRIO EXECUTIVO", title_style))
        
        # Métricas principais
        elements.append(Paragraph("INDICADORES PRINCIPAIS", styles['Heading2']))
        
        # Organizar métricas em tabela
        metrics_table_data = []
        for key, value in metrics.items():
            if key not in ['período', 'gerado_em']:
                metrics_table_data.append([
                    key.replace('_', ' ').title(),
                    str(value)
                ])
        
        metrics_table = Table(metrics_table_data, colWidths=[4*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(metrics_table)
        elements.append(Spacer(1, 30))
        
        # Informações do período
        elements.append(Paragraph(f"<b>Período analisado:</b> {metrics.get('período', 'N/A')}", styles['Normal']))
        elements.append(Paragraph(f"<b>Relatório gerado em:</b> {metrics.get('gerado_em', 'N/A')}", styles['Normal']))
        
        doc.build(elements)
        return buffer.getvalue()

    # Métodos auxiliares
    async def _get_daily_metrics(self, date_from: Optional[date], date_to: Optional[date]) -> List[Dict]:
        """Buscar métricas diárias para gráficos"""
        # Implementação simplificada - pode ser expandida
        daily_data = []
        
        # Por enquanto retorna dados fictícios para demonstração
        # Em produção, faria consultas SQL agrupadas por data
        
        return daily_data

    def _translate_status(self, status: str) -> str:
        """Traduzir status de agendamentos"""
        translations = {
            'pending': 'Pendente',
            'confirmed': 'Confirmado',
            'completed': 'Concluído',
            'cancelled': 'Cancelado'
        }
        return translations.get(status, status)

    def _translate_conversation_status(self, status: str) -> str:
        """Traduzir status de conversas"""
        translations = {
            'active': 'Ativa',
            'closed': 'Encerrada',
            'waiting': 'Aguardando'
        }
        return translations.get(status, status)

    def _calculate_status_breakdown(self, statuses: List[str]) -> Dict:
        """Calcular breakdown de status"""
        from collections import Counter
        counter = Counter(statuses)
        return dict(counter)

    def _calculate_conversation_status_breakdown(self, statuses: List[str]) -> Dict:
        """Calcular breakdown de status de conversas"""
        from collections import Counter
        counter = Counter(statuses)
        return dict(counter)

# Singleton instance
export_service = ReportExportService()
